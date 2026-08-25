"""Builds a synthetic .xlsx in memory that mirrors the real legacy workbook's
layout and quirks (trailing-space headers, a LOCATION="Customer" row with
QTY=0, a datetime-typed date cell, repeated brand/model rows needing product
dedup, a '2nd floor Location' sub-location code) — without committing any
real customer data to the repository (see docs/architecture/09, Prompt 6).
"""

import datetime
import io

import openpyxl

# Deliberately mirrors the real file's header text exactly, including the
# trailing space after "S/N" that the real workbook has.
HEADER_ROW = [
    "BRAND",
    "MODEL/Part No./SKU",
    "TYPE/DESCRIPTION",
    "S/N ",
    "QTY",
    "LOCATION",
    "2nd floor Location",
    "Project Ref. #",
    "FINAL CUSTOMER",
    "COMMENTS/#No",
    "PRODUCT DELIVERY / PRODUCT REMOVAL",
    "Arrival Date",
    "Delivery Date",
    "Return Date",
    "Removal Date",
    "Registrar",
]

DATA_ROWS = [
    # Brand, Model, Type, S/N, QTY, LOCATION, 2nd floor, ProjectRef, Customer, ...
    [
        "Acme Telecom ",
        "AT-D70",
        "Phones",
        "1SYNTH01LF",
        0,
        "Customer",
        None,
        "Q000001",
        "Example Bank",
        None,
    ],
    [
        "Northwind Networks ",
        "NW80W",
        "Firewall",
        "SY21B01742",
        1,
        "Basement 1",
        8,
        "Q9900",
        None,
        None,
    ],
    [
        "Northwind Networks ",
        "NW80W",
        "Firewall",
        "SY21B01813",
        1,
        "Basement 1",
        8,
        "Q9900",
        None,
        None,
    ],
    [
        "Northwind Networks ",
        "NW80W",
        "Firewall",
        "SY21B01719",
        1,
        "Basement 1",
        8,
        "Q9900",
        None,
        None,
    ],
    [
        "Northwind Networks ",
        "NW80W",
        "Firewall",
        "SY22108425",
        1,
        "Basement 1",
        8,
        "Q9900",
        None,
        None,
    ],
    [
        "Contoso Systems",
        "CS881",
        "Router ",
        "SFSY2422C0PK",
        0,
        "Customer",
        None,
        "Q8800",
        "Acme Corp",
        None,
    ],
    [
        "Contoso Systems",
        "CS881",
        "Router ",
        "SFSY2413C362",
        1,
        "Basement 1",
        7,
        "Q8800",
        "Acme Corp",
        "Northtown ",
    ],
]


def build_legacy_workbook_bytes():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(HEADER_ROW)
    for row in DATA_ROWS:
        sheet.append(row)
    # Row 6 (Contoso "Customer" row) carries real datetime cells, matching
    # how openpyxl surfaces a formatted date cell from the real workbook.
    sheet.cell(row=7, column=13, value=datetime.datetime(2025, 12, 17))  # Delivery Date
    sheet.cell(row=7, column=15, value=datetime.datetime(2025, 12, 15))  # Removal Date

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
