from datetime import date

import openpyxl
import pytest
from django.core.exceptions import PermissionDenied, ValidationError

from apps.exports.models import ExportRunStatus, ExportSchedule, ExportSettings
from apps.exports.services import (
    build_inventory_workbook,
    run_export,
    should_run_today,
    update_settings,
)
from apps.inventory.services.receipts import receive_stock


@pytest.mark.django_db
class TestUpdateSettings:
    def test_saves_valid_path_and_schedule(self, administrator, tmp_path):
        settings_obj = update_settings(
            user=administrator,
            export_path=str(tmp_path),
            schedule=ExportSchedule.NIGHTLY,
            weekly_weekday=2,
        )
        assert settings_obj.export_path == str(tmp_path)
        assert settings_obj.schedule == ExportSchedule.NIGHTLY
        assert settings_obj.updated_by == administrator

    def test_rejects_unwritable_path(self, administrator):
        with pytest.raises(ValidationError, match="not writable"):
            update_settings(
                user=administrator,
                export_path="Z:\\definitely\\not\\a\\real\\path",
                schedule=ExportSchedule.NIGHTLY,
                weekly_weekday=6,
            )

    def test_requires_path_when_schedule_enabled(self, administrator):
        with pytest.raises(ValidationError, match="export path is required"):
            update_settings(
                user=administrator, export_path="", schedule=ExportSchedule.WEEKLY, weekly_weekday=6
            )

    def test_disabled_schedule_allows_blank_path(self, administrator):
        settings_obj = update_settings(
            user=administrator, export_path="", schedule=ExportSchedule.DISABLED, weekly_weekday=6
        )
        assert settings_obj.export_path == ""

    def test_requires_administrator(self, stock_manager, tmp_path):
        with pytest.raises(PermissionDenied):
            update_settings(
                user=stock_manager,
                export_path=str(tmp_path),
                schedule=ExportSchedule.NIGHTLY,
                weekly_weekday=6,
            )

    def test_singleton_is_reused(self, administrator, tmp_path):
        update_settings(
            user=administrator,
            export_path=str(tmp_path),
            schedule=ExportSchedule.NIGHTLY,
            weekly_weekday=6,
        )
        assert ExportSettings.objects.count() == 1


@pytest.mark.django_db
class TestBuildInventoryWorkbook:
    def test_includes_units_and_balances(
        self, administrator, unit_product, quantity_product, location_tree
    ):
        receive_stock(
            user=administrator,
            product=unit_product,
            location=location_tree["room"],
            occurred_at=date.today(),
            vendor_serial="SN-EXPORT-1",
        )
        receive_stock(
            user=administrator,
            product=quantity_product,
            location=location_tree["room"],
            occurred_at=date.today(),
            quantity=9,
        )

        workbook = build_inventory_workbook()
        assert workbook.sheetnames == ["Unit Assets", "Stock Balances"]

        assets_sheet = workbook["Unit Assets"]
        serials = [row[4] for row in assets_sheet.iter_rows(min_row=2, values_only=True)]
        assert "SN-EXPORT-1" in serials

        balances_sheet = workbook["Stock Balances"]
        rows = list(balances_sheet.iter_rows(min_row=2, values_only=True))
        assert any(row[5] == 9 for row in rows)


@pytest.mark.django_db
class TestRunExport:
    def test_writes_a_real_readable_workbook(
        self, administrator, unit_product, location_tree, tmp_path
    ):
        receive_stock(
            user=administrator,
            product=unit_product,
            location=location_tree["room"],
            occurred_at=date.today(),
            vendor_serial="SN-RUNEXPORT-1",
        )
        update_settings(
            user=administrator,
            export_path=str(tmp_path),
            schedule=ExportSchedule.NIGHTLY,
            weekly_weekday=6,
        )

        path = run_export(user=administrator)

        assert path.startswith(str(tmp_path))
        workbook = openpyxl.load_workbook(path)
        serials = [row[4] for row in workbook["Unit Assets"].iter_rows(min_row=2, values_only=True)]
        assert "SN-RUNEXPORT-1" in serials

        settings_obj = ExportSettings.load()
        assert settings_obj.last_run_status == ExportRunStatus.SUCCESS
        assert settings_obj.last_run_at is not None

    def test_records_failure_status_on_unwritable_path(self, administrator, tmp_path):
        update_settings(
            user=administrator,
            export_path=str(tmp_path),
            schedule=ExportSchedule.NIGHTLY,
            weekly_weekday=6,
        )
        # Bypass form-level validation to simulate a path that was reachable
        # when saved but has since become unreachable (e.g. a disconnected
        # network share) — the real-world case run_export()'s own error
        # handling exists for.
        settings_obj = ExportSettings.load()
        settings_obj.export_path = "Z:\\now\\unreachable"
        settings_obj.save(update_fields=["export_path"])

        with pytest.raises(OSError):
            run_export(user=administrator)

        settings_obj.refresh_from_db()
        assert settings_obj.last_run_status == ExportRunStatus.FAILED
        assert settings_obj.last_run_detail

    def test_raises_when_no_path_configured(self, administrator):
        with pytest.raises(ValidationError, match="No export path"):
            run_export(user=administrator)


@pytest.mark.django_db
class TestShouldRunToday:
    def test_disabled_never_runs(self):
        settings_obj = ExportSettings(schedule=ExportSchedule.DISABLED)
        assert should_run_today(settings_obj) is False

    def test_nightly_always_runs(self):
        settings_obj = ExportSettings(schedule=ExportSchedule.NIGHTLY)
        assert should_run_today(settings_obj, today=date(2026, 1, 1)) is True
        assert should_run_today(settings_obj, today=date(2026, 1, 5)) is True

    def test_weekly_only_runs_on_configured_day(self):
        settings_obj = ExportSettings(schedule=ExportSchedule.WEEKLY, weekly_weekday=6)
        sunday = date(2026, 8, 30)
        monday = date(2026, 8, 31)
        assert sunday.weekday() == 6
        assert should_run_today(settings_obj, today=sunday) is True
        assert should_run_today(settings_obj, today=monday) is False
